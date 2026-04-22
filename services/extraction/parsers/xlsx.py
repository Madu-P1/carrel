from __future__ import annotations

from pathlib import Path

from fastapi import HTTPException

from ..types import ExtractedElement
from ..utils import safe_text, file_sha
from .common import ParserContext, build_asset, make_span

try:
    import openpyxl
except ImportError:
    openpyxl = None


def parse_xlsx(path: Path, *, suffix: str, mime_type: str, context: ParserContext):
    if openpyxl is None:
        raise HTTPException(status_code=400, detail="XLSX support requires openpyxl")
    wb_values = openpyxl.load_workbook(str(path), data_only=True)
    wb_formulas = openpyxl.load_workbook(str(path), data_only=False)
    file_id = file_sha(path)[:16]
    elements: list[ExtractedElement] = []
    warnings: list[str] = []
    try:
        for sheet_name in wb_values.sheetnames:
            ws_values = wb_values[sheet_name]
            ws_formulas = wb_formulas[sheet_name]
            span = make_span(path, file_id, sheet=sheet_name, section=sheet_name, element_id=f"sheet-{sheet_name}")
            elements.append(
                ExtractedElement(
                    id=span.element_id or f"sheet-{sheet_name}",
                    kind="sheet",
                    text=sheet_name,
                    normalized_text=sheet_name,
                    span=span,
                )
            )
            merged_ranges = [str(item) for item in getattr(ws_formulas.merged_cells, "ranges", [])]
            if merged_ranges:
                warnings.append(f"Sheet {sheet_name} contains merged ranges: {', '.join(merged_ranges[:4])}")
            for row_index in range(1, ws_values.max_row + 1):
                display_values: list[str] = []
                cell_meta: list[dict[str, object]] = []
                for column_index in range(1, ws_values.max_column + 1):
                    formula_cell = ws_formulas.cell(row=row_index, column=column_index)
                    display_cell = ws_values.cell(row=row_index, column=column_index)
                    display = display_cell.value
                    formula = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None
                    if display is None and formula is None:
                        continue
                    display_text = safe_text(display if display is not None else formula)
                    if not display_text:
                        continue
                    coord = formula_cell.coordinate
                    display_values.append(f"{coord}: {display_text}")
                    cell_meta.append({"coordinate": coord, "display": display, "formula": formula})
                    span = make_span(
                        path,
                        file_id,
                        sheet=sheet_name,
                        section=sheet_name,
                        row_range=f"{row_index}:{row_index}",
                        cell_range=coord,
                        element_id=f"{sheet_name}-{coord}",
                    )
                    elements.append(
                        ExtractedElement(
                            id=span.element_id or f"{sheet_name}-{coord}",
                            kind="cell",
                            text=display_text,
                            normalized_text=display_text,
                            span=span,
                            metadata={"formula": formula, "coordinate": coord},
                        )
                    )
                if display_values:
                    row_text = "\n".join(display_values)
                    span = make_span(
                        path,
                        file_id,
                        sheet=sheet_name,
                        section=sheet_name,
                        row_range=f"{row_index}:{row_index}",
                        element_id=f"{sheet_name}-row-{row_index}",
                    )
                    elements.append(
                        ExtractedElement(
                            id=span.element_id or f"{sheet_name}-row-{row_index}",
                            kind="table_row",
                            text=row_text,
                            normalized_text=row_text,
                            span=span,
                            metadata={"cells": cell_meta},
                        )
                    )
    finally:
        wb_values.close()
        wb_formulas.close()
    return build_asset(
        path,
        detected_type=suffix,
        mime_type=mime_type,
        parser_name="openpyxl-structured",
        elements=elements,
        context=context,
        warnings=warnings,
        extraction_modes=["table_aware"],
        metadata={"page_count": len(wb_values.sheetnames)},
        confidence=0.82,
    )
